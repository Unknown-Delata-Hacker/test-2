import streamlit as st
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
import schedule
import time
path = 'persons22'
images = []
classNames = []
myList = os.listdir(path)
ids = []
users = []
############################################### read images and spliting names and ids each ####################################
for i in myList:
    # Split the filename into name and extension
    name, ext = os.path.splitext(i)
    
    # Extract the number from the filename after the dot
    id = name.split('.')[-1]  # Get the part after the last dot as the "id"
    user = name.split('.')[0]
    currntImg = cv2.imread(f'{path}/{i}')
    images.append(currntImg)
    users.append(user)
    classNames.append(name)  # Add the filename (without extension) to classNames list
    ids.append(id)  # Add the extracted id to the ids list

# print(classNames)
print(ids)
print(users)
############################################### Encoding the images  ####################################
def findEncodings(images):
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList
############################################### Marking Attendance With Excel   ####################################
# import openpyxl
# def markAttendance(name):
#     with open('Attendance.xlsx','r+') as f:
#         myDataList = f.readlines()
#         nameList = []
#         for line in myDataList:
#             entry = line.split(',')
#             nameList.append(entry[0])
#         if name not in nameList:                                      ###  old code 
#             now = datetime.now()
#             dtString = now.strftime('%H:%M:%S')
#             f.writelines(f'\n{name},{dtString}')
        

# encodeListKnown = findEncodings(images)
# print('Encoding Complete')

import openpyxl
from openpyxl import Workbook
from datetime import date, datetime

def make_attendance_file(sheet_name):
    wb = Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=1).value = "Names"
    sheet.cell(row=1, column=2).value = "Time"
    sheet.cell(row=1, column=3).value = "ID"
    wb.save(filename=f'{sheet_name}.xlsx')

def make_attendance(name,sheet_name):
    filename = f'{sheet_name}.xlsx'
    if not os.path.isfile (filename):
        make_attendance_file(sheet_name)
    
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    id = name.split('.')[-1]  # Get the part after the last dot as the "id"
    user = name.split('.')[0]

    for row in range(2, sheet.max_row + 1):
        if user == str(sheet.cell(row, 1).value):
            # Write the current time
            sheet.cell(row=row, column=2).value = datetime.now().strftime("%H:%M:%S")
            break
    else:
        # Add the name and the time if not found

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = user
        sheet.cell(row=next_row, column=2).value = datetime.now().strftime("%H:%M:%S")
        sheet.cell(row=next_row, column=3).value = id

    wb.save(filename=filename)
encodeListKnown = findEncodings(images)
print('Encoding Complete')

############################################### Camera Access  ####################################
cap = cv2.VideoCapture(0)
frame_resizing = 0.25

while True:
    success, img = cap.read()
    imgS = cv2.resize(img, (0, 0), fx=frame_resizing, fy=frame_resizing)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
 
    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS,facesCurFrame)
 
    for encodeFace,faceLoc in zip(encodesCurFrame,facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown,encodeFace)
        print('matches',matches)
        faceDis = face_recognition.face_distance(encodeListKnown,encodeFace)
        print(faceDis)
        matchIndex = np.argmin(faceDis)
        if matches[matchIndex]:
            name = classNames[matchIndex].upper()
            print(name)
            
            faceLoc = np.array(faceLoc)
            faceLoc = faceLoc / 0.25
            faceLoc=faceLoc.astype(int)
            #y1,x2,y2,x1 = faceLoc # in the other code we risize the face
            #y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
            y1, x2, y2, x1 = faceLoc[0], faceLoc[1], faceLoc[2], faceLoc[3]
         
            cv2.rectangle(img,(x1,y1),(x2,y2),(0,255,0),2)
            cv2.rectangle(img,(x1,y2-35),(x2,y2),(0,255,0),cv2.FILLED)
            cv2.putText(img,name,(x1+6,y2-6),cv2.FONT_HERSHEY_COMPLEX,1,(255,255,255),2)
            make_attendance(name,"Arabic")
 
    cv2.imshow('Webcam',img)
    key = cv2.waitKey(1)
    if key == 27:
        break
cap.release()
cv2.destroyAllWindows()

###############################################
###############################################
###############################################     subject Excel sheets 
###############################################
import openpyxl
import os
import schedule
import time
from datetime import datetime, timedelta

def create_excel(subject):
    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # Add some data to the worksheet (for demonstration purposes)
    worksheet['A1'] = f"Welcome to {subject} Class!"
    
    # Save the Excel file with the subject name
    filename = f"{subject}_Class.xlsx"
    workbook.save(filename)
    print(f"Excel file '{filename}' created.")

def main():
    st.write("Starting attendance system...")

    subjects = ['Arabic', 'English', 'Math', 'Social']
    
    # Set the start time to 9:00 AM
    start_time = datetime.now().replace(hour=9, minute=0, second=0, microsecond=0)
    
    # Schedule the task every 2 hours within the specified time range (9:00 AM to 5:00 PM)
    for i, subject in enumerate(subjects):
        scheduled_time = start_time + timedelta(hours=2*i)
        if scheduled_time < datetime.now():
            continue  # Skip past schedules
        schedule.every().day.at(scheduled_time.strftime("%H:%M")).do(create_excel, subject)
    
    # Run the schedule loop
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == "__main__":
    st.title("Attendance System")

    st.button("Start Attendance System", on_click=main)
    st.button("End Attendance System", on_click=quit)
